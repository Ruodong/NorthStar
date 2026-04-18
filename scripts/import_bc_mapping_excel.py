#!/usr/bin/env python3
"""Import Business Capability → Application mapping from an Excel file
into northstar.ref_app_business_capability.

This is a host-side, one-off / ad-hoc tool for when an architect sends
us an XLSX sheet of mappings that haven't been (or won't be) round-tripped
through the EAM source system. For regular EAM-sourced mapping updates,
use scripts/sync_from_egm.py instead.

## What the script does

1. Read mapping rows from the Excel (default sheet: "EA Management")
2. For each row, look up `bcpf_master_id` in `ref_business_capability`
   by (bc_id, data_version). This requires the BC master to already be
   synced — run `sync_from_egm.py --only ref_business_capability` first.
3. Skip rows where:
   - The BC master lookup fails (bc_id+version not present)
   - The (app_id, bcpf_master_id) pair already exists in the mapping
     table (from EAM sync or a prior import)
4. INSERT the rest with a random UUIDv4 id. No-op on re-run except for
   rows that are genuinely new.

## Usage

    # Dry-run (default): print what would happen, don't modify DB
    python3 scripts/import_bc_mapping_excel.py \\
        --file "data/Business Capability Mapping 20260418.xlsx"

    # Execute: actually insert
    python3 scripts/import_bc_mapping_excel.py \\
        --file "data/Business Capability Mapping 20260418.xlsx" \\
        --execute

The DB target is the NorthStar postgres pointed at by these env vars
(same defaults as sync_from_egm.py):
    NORTHSTAR_PG_HOST     default localhost
    NORTHSTAR_PG_PORT     default 5434
    NORTHSTAR_PG_DB       default northstar
    NORTHSTAR_PG_USER     default northstar
    POSTGRES_PASSWORD     default northstar_dev
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
import uuid
from pathlib import Path

import openpyxl
import psycopg

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s - %(message)s"
)
logger = logging.getLogger("bc-import")


# Column indexes in the "EA Management" sheet (header row 0, data row 1+).
# Kept as named constants so layout changes are a one-line fix.
COL_APP_ID = 0
COL_BC_ID = 12
COL_LEVEL = 15
COL_VERSION = 20
COL_CREATED_AT = 21
COL_CREATED_BY = 22


def northstar_dsn() -> str:
    return (
        f"host={os.environ.get('NORTHSTAR_PG_HOST', 'localhost')} "
        f"port={os.environ.get('NORTHSTAR_PG_PORT', '5434')} "
        f"dbname={os.environ.get('NORTHSTAR_PG_DB', 'northstar')} "
        f"user={os.environ.get('NORTHSTAR_PG_USER', 'northstar')} "
        f"password={os.environ.get('POSTGRES_PASSWORD', 'northstar_dev')}"
    )


def read_excel_rows(path: Path, sheet: str) -> list[dict]:
    """Load mapping rows from the spreadsheet as a list of dicts.

    Only rows where BC ID is populated are returned. Whitespace is
    trimmed from string values.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {sheet!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)  # skip header

    out: list[dict] = []
    for r in rows_iter:
        app_id = r[COL_APP_ID]
        bc_id = r[COL_BC_ID]
        if not app_id or not bc_id:
            continue
        version = r[COL_VERSION]
        # Normalize version — Excel gives floats like 1.4, DB stores "1.4"
        if isinstance(version, float):
            version_str = f"{version:.1f}".rstrip("0").rstrip(".")
            if "." not in version_str:
                version_str += ".0"
        else:
            version_str = str(version) if version is not None else None
        out.append({
            "app_id": str(app_id).strip(),
            "bc_id": str(bc_id).strip(),
            "data_version": version_str,
            "level": r[COL_LEVEL],
            "source_created_at": r[COL_CREATED_AT],
            "source_create_by": r[COL_CREATED_BY],
        })
    return out


def build_bc_lookup(cur) -> dict[tuple[str, str], int]:
    """(bc_id, data_version) → bcpf_master_id (bigint).

    The Excel's `Versions` column is the taxonomy version the mapping
    was authored against, so we must match on both bc_id AND version
    (bc_id alone isn't unique across taxonomy versions).
    """
    cur.execute(
        "SELECT bc_id, data_version, id FROM northstar.ref_business_capability"
    )
    out: dict[tuple[str, str], int] = {}
    for bc_id, ver, master_id in cur:
        out[(bc_id, ver)] = master_id
    return out


def load_existing_pairs(cur) -> set[tuple[str, int]]:
    """(app_id, bcpf_master_id) pairs already in the mapping table.

    Used to skip duplicates on re-import. Catches both EAM-synced rows
    and rows from a prior run of this script.
    """
    cur.execute(
        "SELECT app_id, bcpf_master_id FROM northstar.ref_app_business_capability"
    )
    return set((r[0], r[1]) for r in cur)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--file",
        required=True,
        help="Path to the mapping Excel (e.g. data/Business Capability Mapping 20260418.xlsx)",
    )
    ap.add_argument(
        "--sheet",
        default="EA Management",
        help='Sheet name (default: "EA Management")',
    )
    ap.add_argument(
        "--execute",
        action="store_true",
        help="Actually insert rows. Default is dry-run (no DB changes).",
    )
    args = ap.parse_args()

    path = Path(args.file)
    if not path.is_file():
        logger.error("File not found: %s", path)
        return 2

    excel_rows = read_excel_rows(path, args.sheet)
    logger.info("Read %d mapping rows from %s / %s", len(excel_rows), path.name, args.sheet)

    if not excel_rows:
        logger.info("Nothing to do.")
        return 0

    dsn = northstar_dsn()
    logger.info(
        "NorthStar DSN: %s:%s/%s",
        os.environ.get("NORTHSTAR_PG_HOST", "localhost"),
        os.environ.get("NORTHSTAR_PG_PORT", "5434"),
        os.environ.get("NORTHSTAR_PG_DB", "northstar"),
    )

    with psycopg.connect(dsn) as conn:
        with conn.cursor() as cur:
            bc_lookup = build_bc_lookup(cur)
            existing_pairs = load_existing_pairs(cur)
            logger.info(
                "Loaded %d master BCs; %d mapping pairs already present",
                len(bc_lookup),
                len(existing_pairs),
            )

            to_insert: list[dict] = []
            skipped_dup = 0
            skipped_no_master = 0
            no_master_samples: list[tuple[str, str]] = []

            for row in excel_rows:
                key = (row["bc_id"], row["data_version"])
                master_id = bc_lookup.get(key)
                if master_id is None:
                    skipped_no_master += 1
                    if len(no_master_samples) < 10:
                        no_master_samples.append(key)
                    continue
                pair = (row["app_id"], master_id)
                if pair in existing_pairs:
                    skipped_dup += 1
                    continue
                to_insert.append({
                    **row,
                    "bcpf_master_id": master_id,
                    "id": str(uuid.uuid4()),
                })
                # Prevent duplicate inserts within this batch if Excel
                # has the same (app, bc) twice somehow
                existing_pairs.add(pair)

            logger.info("=== PLAN ===")
            logger.info("Will INSERT:         %d rows", len(to_insert))
            logger.info("Skip (dup in DB):    %d rows", skipped_dup)
            logger.info("Skip (no master):    %d rows", skipped_no_master)
            if no_master_samples:
                logger.info(
                    "  sample no-master (bc_id, version): %s",
                    no_master_samples,
                )

            if not args.execute:
                logger.info("DRY RUN. Pass --execute to actually insert.")
                return 0

            if not to_insert:
                logger.info("Nothing new to insert.")
                return 0

            insert_sql = """
                INSERT INTO northstar.ref_app_business_capability
                    (id, app_id, bcpf_master_id, bc_id, data_version,
                     source_create_by, source_created_at, synced_at)
                VALUES
                    (%s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (id) DO NOTHING
            """
            cur.executemany(
                insert_sql,
                [
                    (
                        r["id"],
                        r["app_id"],
                        r["bcpf_master_id"],
                        r["bc_id"],
                        r["data_version"],
                        r["source_create_by"],
                        r["source_created_at"],
                    )
                    for r in to_insert
                ],
            )
            conn.commit()
            logger.info("INSERTED %d new mapping rows.", len(to_insert))

    return 0


if __name__ == "__main__":
    sys.exit(main())
