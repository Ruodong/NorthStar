#!/usr/bin/env python3
"""Load integration_catalog.xlsx into northstar.integration_interface.

This replaces the drawio-extracted confluence_diagram_interaction as the
authoritative source of application integration edges.

Source: data/integration_catalog.xlsx (10 sheets, ~30k rows)
Target: northstar.integration_interface

Behavior:
  - Handles Excel merged cells (anchor value fills the merged range).
    Critical for APIH (1999 merges) and KPaaS (10789 merges) where Pub-side
    CMDB IDs span multiple rows.
  - Per-platform column mapping: each sheet has a different header layout;
    we project all into the unified integration_interface schema.
  - Name-based CMDB resolution: rows missing source_cmdb_id / target_cmdb_id
    but with source_app_name / target_app_name get trigram-matched against
    ref_application. Only high-confidence matches (similarity >= 0.65) are
    accepted to avoid false positives.
  - Idempotent: source_row_hash = sha256 of canonicalized field set.
    Re-running overwrites on UNIQUE (integration_platform, source_row_hash).

Usage (from ~/NorthStar on 71):
    set -a && source .env && set +a
    .venv-ingest/bin/python scripts/load_integration_catalog.py \
        --file data/integration_catalog.xlsx [--dry-run]
"""
from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path as _P
from typing import Any, Optional

import openpyxl
import psycopg
from psycopg.rows import dict_row

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s - %(message)s",
)
logger = logging.getLogger("load-int-catalog")


# -----------------------------------------------------------------------------
# DB connection
# -----------------------------------------------------------------------------
def pg_dsn() -> str:
    return (
        f"host={os.environ.get('NORTHSTAR_PG_HOST', 'localhost')} "
        f"port={os.environ.get('NORTHSTAR_PG_PORT', '5434')} "
        f"dbname={os.environ.get('NORTHSTAR_PG_DB', 'northstar')} "
        f"user={os.environ.get('NORTHSTAR_PG_USER', 'northstar')} "
        f"password={os.environ.get('POSTGRES_PASSWORD', 'northstar_dev')}"
    )


# -----------------------------------------------------------------------------
# Excel loading with merged cell resolution
# -----------------------------------------------------------------------------
def load_sheet_with_merges(ws) -> list[list[Any]]:
    """Read a worksheet into a 2D array with merged cells filled.

    Merged ranges have the value only in the top-left anchor cell when read
    directly; all other cells in the range return None. We fill the entire
    merge range with the anchor value so downstream code sees intuitive rows.
    """
    max_row, max_col = ws.max_row, ws.max_column
    grid: list[list[Any]] = [[None] * max_col for _ in range(max_row)]
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        for cell in row:
            grid[cell.row - 1][cell.column - 1] = cell.value
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = mr.min_col, mr.min_row, mr.max_col, mr.max_row
        anchor = grid[min_r - 1][min_c - 1]
        for r in range(min_r - 1, max_r):
            for c in range(min_c - 1, max_c):
                grid[r][c] = anchor
    return grid


# -----------------------------------------------------------------------------
# Per-platform column mappings
# -----------------------------------------------------------------------------
# Each mapping is a dict of target_column → list of candidate source headers
# (the first header present in the sheet wins). This lets us handle the
# heterogeneous sheet layouts with one table.

COMMON_MAP: dict[str, list[str]] = {
    "interface_name":          ["interface_name"],
    "source_cmdb_id":          ["source_cmdb_id"],
    "target_cmdb_id":          ["target_cmdb_id"],
    "source_app_name":         ["s_application_name", "源系统名称"],
    "target_app_name":         ["t_application_name", "目标源系统名称"],
    "source_endpoint":         ["source_endpoint"],
    "target_endpoint":         ["target_endpoint"],
    "source_version":          ["source_version"],
    "target_version":          ["target_version"],
    "source_dc":               ["source_dc"],
    "target_dc":               ["target_dc"],
    "source_application_type": ["source_application_type"],
    "target_application_type": ["target_application_type"],
    "source_connection_type":  ["source_connection_type"],
    "target_connection_type":  ["target_connection_type"],
    "source_authentication":   ["source_authentication"],
    "target_authentication":   ["target_authentication"],
    "interface_owner":         ["interface_owner"],
    "source_owner":            ["source_owner"],
    "target_owner":            ["target_owner"],
    "s_team_publicmail":       ["s_team_publicmail"],
    "t_team_publicmail":       ["t_team_publicmail"],
    "s_application_linemanager": ["s_application_linemanager"],
    "t_application_linemanager": ["t_application_linemanager"],
    "developer":               ["developer"],
    "frequency":               ["frequency"],
    "schedule":                ["schedule"],
    "status":                  ["status"],
    "business_area":           ["business_area"],
    "interface_description":   ["interface_description", "Description"],
    "location":                ["location", "PO Server地址"],
    "api_postman_url":         ["api_postman_url"],
    "api_spec":                ["api_spec"],
    "api_payload_size":        ["api_payload_size"],
    "source_payload_size":     ["source_payload_size"],
    "target_payload_size":     ["target_payload_size"],
    "data_mapping_file":       ["data_mapping_file"],
    "base":                    ["base"],
    "git_project":             ["git_project"],
    "version":                 ["version"],
    "tag":                     ["tag"],
}

# Fields that go into raw_fields jsonb instead of typed columns (the long tail)
RAW_FIELDS_KEEP = {
    "performance_report",
    "source_trace_field",
    "target_trace_field",
    "source_application_inbound_sample",
    "source_application_outbound_sample",
    "target_application_inbound_sample",
    "target_application_outbound_sample",
    "return_code_description",
}

# APIH sheet has different column names entirely — publisher/subscriber model
APIH_MAP: dict[str, str] = {
    "source_cmdb_id":     "Pub CMDB ID",
    "target_cmdb_id":     "Sub Cmdb",
    "source_app_name":    "Pub CMDB Name",
    "target_app_name":    "Sub CMDB Name",
    "source_account_name": "Pub Account Name",
    "target_account_name": "Sub Account Name",
    "api_name":           "Pub API Name",
    "instance":           "Instance",
    "interface_description": "Description",
}

# KPaaS is Kafka — publisher topic / subscriber
KPAAS_MAP: dict[str, str] = {
    "source_cmdb_id":     "Pub CMDB ID",
    "target_cmdb_id":     "Sub CMDB",
    "source_app_name":    "Pub App Name",
    "target_app_name":    "Sub App Name",
    "source_account_name": "Pub Account Name",
    "target_account_name": "Sub Account Name",
    "topic_name":         "Pub Topic Name",
    "instance":           "Instance",
    "interface_description": "Description",
}


# -----------------------------------------------------------------------------
# Row canonicalization + hash
# -----------------------------------------------------------------------------
def canonicalize_row(d: dict[str, Any]) -> str:
    """Deterministic hash of the non-null fields in this row.

    Used as idempotency key in UNIQUE(integration_platform, source_row_hash).
    Two reingests of the same source row produce the same hash, so we
    UPDATE instead of INSERT-duplicate.
    """
    canonical = {
        k: str(v).strip()
        for k, v in d.items()
        if v is not None and str(v).strip() != ""
    }
    canon_str = json.dumps(canonical, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(canon_str.encode("utf-8")).hexdigest()[:32]


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# -----------------------------------------------------------------------------
# Per-sheet row extractor
# -----------------------------------------------------------------------------
def extract_row(
    platform: str,
    headers: dict[str, int],
    row: list[Any],
) -> dict[str, Any]:
    """Map a single sheet row to the unified schema dict."""
    out: dict[str, Any] = {"integration_platform": platform}

    # APIH uses entirely different headers
    if platform == "APIH":
        for tgt, src in APIH_MAP.items():
            if src in headers:
                out[tgt] = _norm(row[headers[src]])
        # Instance doubles as "location" for APIH
        if out.get("instance") and "location" not in out:
            out["location"] = out["instance"]
        return out

    if platform == "KPaaS":
        for tgt, src in KPAAS_MAP.items():
            if src in headers:
                out[tgt] = _norm(row[headers[src]])
        if out.get("instance") and "location" not in out:
            out["location"] = out["instance"]
        # KPaaS is inherently Kafka — mark connection type explicitly
        out.setdefault("source_connection_type", "Kafka")
        out.setdefault("target_connection_type", "Kafka")
        return out

    # All other sheets share COMMON_MAP
    for tgt, candidates in COMMON_MAP.items():
        for src in candidates:
            if src in headers:
                v = _norm(row[headers[src]])
                if v is not None:
                    out[tgt] = v
                    break

    # Raw fields: stash long-tail platform-specific columns
    raw: dict[str, Any] = {}
    for src_name, idx in headers.items():
        if src_name in RAW_FIELDS_KEEP:
            v = _norm(row[idx])
            if v is not None:
                raw[src_name] = v
    if raw:
        out["raw_fields"] = raw

    return out


# -----------------------------------------------------------------------------
# CMDB resolution — build name → app_id map from ref_application
# -----------------------------------------------------------------------------
def load_cmdb_apps(pg: psycopg.Connection) -> tuple[set[str], dict[str, str]]:
    """Return (set of valid A-ids, dict name_lower → app_id).

    The name dict is best-effort exact (case-insensitive) matching. Trigram
    fuzzy matching happens later via SQL for unresolved rows.
    """
    valid_ids: set[str] = set()
    name_to_id: dict[str, str] = {}
    with pg.cursor() as cur:
        cur.execute("SELECT app_id, name, app_full_name FROM northstar.ref_application")
        for row in cur.fetchall():
            valid_ids.add(row["app_id"])
            for name in (row.get("name"), row.get("app_full_name")):
                if name:
                    key = str(name).strip().lower()
                    # First writer wins; skip if collision (ambiguous name)
                    if key and key not in name_to_id:
                        name_to_id[key] = row["app_id"]
    logger.info(
        "loaded CMDB: %d apps, %d unique (case-insensitive) names",
        len(valid_ids), len(name_to_id),
    )
    return valid_ids, name_to_id


# -----------------------------------------------------------------------------
# Account-name reverse-lookup map (for APIH / KPaaS)
# -----------------------------------------------------------------------------
def build_account_map(
    all_rows: list[dict[str, Any]],
) -> dict[str, str]:
    """Build `account_name → cmdb_id` map from rows where both are present.

    The user pointed out that APIH/KPaaS merged cells mostly solve the Pub-side
    gap. But where Pub still leaks (e.g., Pub Account Name filled but Pub CMDB
    ID not covered by merge), we can reverse-lookup via Sub side entries.
    """
    m: dict[str, str] = {}
    for r in all_rows:
        # Source side
        acc = r.get("source_account_name")
        cid = r.get("source_cmdb_id")
        if acc and cid and acc not in m:
            m[acc] = cid
        # Target side
        acc = r.get("target_account_name")
        cid = r.get("target_cmdb_id")
        if acc and cid and acc not in m:
            m[acc] = cid
    return m


# -----------------------------------------------------------------------------
# CMDB resolution passes
# -----------------------------------------------------------------------------
def resolve_cmdb_ids(
    rows: list[dict[str, Any]],
    valid_ids: set[str],
    name_to_id: dict[str, str],
    account_map: dict[str, str],
    stats: dict[str, int],
) -> None:
    """Mutate rows in-place: fill source_cmdb_id / target_cmdb_id where missing.

    Resolution priority:
      1. direct A-id (already present and valid)
      2. account_name reverse lookup
      3. exact case-insensitive name match against ref_application
      4. leave null (will be retried via trigram in a later SQL pass)
    """
    for r in rows:
        for side in ("source", "target"):
            id_key = f"{side}_cmdb_id"
            name_key = f"{side}_app_name"
            acc_key = f"{side}_account_name"

            current = r.get(id_key)
            if current and current in valid_ids:
                stats["resolved_direct"] += 1
                continue

            # Reject invalid A-ids (not in CMDB) — blank them
            if current and current not in valid_ids:
                r[id_key] = None
                stats["rejected_invalid_id"] += 1

            # (2) account-name reverse lookup
            acc = r.get(acc_key)
            if acc and acc in account_map:
                r[id_key] = account_map[acc]
                stats["resolved_account_lookup"] += 1
                continue

            # (3) exact name match
            name = r.get(name_key)
            if name:
                key = name.strip().lower()
                if key in name_to_id:
                    r[id_key] = name_to_id[key]
                    stats["resolved_name_exact"] += 1
                    continue

            # Fall through — leaves None, handled by trigram SQL later
            stats["unresolved"] += 1


# -----------------------------------------------------------------------------
# Post-ingest trigram fallback (SQL-side)
# -----------------------------------------------------------------------------
TRIGRAM_FILL_SQL = """
WITH candidates AS (
    SELECT
        i.interface_id,
        '{side}'::text AS side,
        i.{side}_app_name AS name,
        (SELECT a.app_id
         FROM northstar.ref_application a
         WHERE a.name IS NOT NULL
         ORDER BY similarity(lower(a.name), lower(i.{side}_app_name)) DESC
         LIMIT 1) AS best_match,
        (SELECT similarity(lower(a.name), lower(i.{side}_app_name))
         FROM northstar.ref_application a
         WHERE a.name IS NOT NULL
         ORDER BY similarity(lower(a.name), lower(i.{side}_app_name)) DESC
         LIMIT 1) AS score
    FROM northstar.integration_interface i
    WHERE i.{side}_cmdb_id IS NULL
      AND i.{side}_app_name IS NOT NULL
      AND length(i.{side}_app_name) >= 3
)
UPDATE northstar.integration_interface i
SET {side}_cmdb_id = c.best_match
FROM candidates c
WHERE i.interface_id = c.interface_id
  AND c.score >= 0.65
  AND c.best_match IS NOT NULL;
"""


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--file",
        default=str(_P(__file__).parent.parent / "data" / "integration_catalog.xlsx"),
    )
    ap.add_argument("--dry-run", action="store_true", help="Parse + resolve, no writes")
    ap.add_argument(
        "--wipe",
        action="store_true",
        help="Truncate integration_interface before inserting",
    )
    args = ap.parse_args()

    xlsx_path = _P(args.file)
    if not xlsx_path.exists():
        logger.error("file not found: %s", xlsx_path)
        return 1

    logger.info("opening %s", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    stats: dict[str, int] = {
        "rows_seen": 0,
        "rows_skipped_empty": 0,
        "rows_ingested": 0,
        "resolved_direct": 0,
        "resolved_account_lookup": 0,
        "resolved_name_exact": 0,
        "rejected_invalid_id": 0,
        "unresolved": 0,
    }

    # Pass 1: read all sheets, extract rows
    all_rows: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid = load_sheet_with_merges(ws)
        if len(grid) < 2:
            continue
        headers = {h: idx for idx, h in enumerate(grid[0]) if h}

        n = 0
        for row_num, row in enumerate(grid[1:], start=2):
            if all(v is None for v in row):
                stats["rows_skipped_empty"] += 1
                continue
            stats["rows_seen"] += 1
            extracted = extract_row(sheet_name, headers, row)
            extracted["source_row_num"] = row_num
            extracted["source_row_hash"] = canonicalize_row({
                k: v for k, v in extracted.items()
                if k not in ("source_row_num",)
            })
            all_rows.append(extracted)
            n += 1
        logger.info("  %s: %d rows", sheet_name, n)

    logger.info("total rows extracted: %d", len(all_rows))

    # Pass 2: build reference maps
    pg = psycopg.connect(pg_dsn(), row_factory=dict_row)
    try:
        valid_ids, name_to_id = load_cmdb_apps(pg)
        account_map = build_account_map(all_rows)
        logger.info("account_name map size: %d", len(account_map))

        # Pass 3: resolve CMDB IDs
        resolve_cmdb_ids(all_rows, valid_ids, name_to_id, account_map, stats)

        if args.dry_run:
            logger.info("dry-run — skipping writes")
            for k, v in stats.items():
                logger.info("  %-30s %d", k, v)
            return 0

        # Pass 4: write to PG
        with pg.cursor() as cur:
            if args.wipe:
                logger.info("wiping integration_interface...")
                cur.execute("TRUNCATE TABLE northstar.integration_interface RESTART IDENTITY")

            insert_sql = """
                INSERT INTO northstar.integration_interface (
                    integration_platform, interface_name, source_row_hash, source_row_num,
                    source_cmdb_id, target_cmdb_id, source_app_name, target_app_name,
                    source_account_name, target_account_name,
                    source_endpoint, target_endpoint, source_version, target_version,
                    source_dc, target_dc,
                    source_application_type, target_application_type,
                    source_connection_type, target_connection_type,
                    source_authentication, target_authentication,
                    interface_owner, source_owner, target_owner,
                    s_team_publicmail, t_team_publicmail,
                    s_application_linemanager, t_application_linemanager,
                    developer,
                    frequency, schedule, status, business_area,
                    interface_description, location,
                    api_name, topic_name, instance, api_postman_url, api_spec,
                    api_payload_size, source_payload_size, target_payload_size,
                    data_mapping_file, base, git_project, version, tag,
                    raw_fields, updated_at
                )
                VALUES (
                    %(integration_platform)s, %(interface_name)s, %(source_row_hash)s, %(source_row_num)s,
                    %(source_cmdb_id)s, %(target_cmdb_id)s, %(source_app_name)s, %(target_app_name)s,
                    %(source_account_name)s, %(target_account_name)s,
                    %(source_endpoint)s, %(target_endpoint)s, %(source_version)s, %(target_version)s,
                    %(source_dc)s, %(target_dc)s,
                    %(source_application_type)s, %(target_application_type)s,
                    %(source_connection_type)s, %(target_connection_type)s,
                    %(source_authentication)s, %(target_authentication)s,
                    %(interface_owner)s, %(source_owner)s, %(target_owner)s,
                    %(s_team_publicmail)s, %(t_team_publicmail)s,
                    %(s_application_linemanager)s, %(t_application_linemanager)s,
                    %(developer)s,
                    %(frequency)s, %(schedule)s, %(status)s, %(business_area)s,
                    %(interface_description)s, %(location)s,
                    %(api_name)s, %(topic_name)s, %(instance)s, %(api_postman_url)s, %(api_spec)s,
                    %(api_payload_size)s, %(source_payload_size)s, %(target_payload_size)s,
                    %(data_mapping_file)s, %(base)s, %(git_project)s, %(version)s, %(tag)s,
                    %(raw_fields)s, NOW()
                )
                ON CONFLICT (integration_platform, source_row_hash) DO UPDATE SET
                    interface_name = EXCLUDED.interface_name,
                    source_cmdb_id = EXCLUDED.source_cmdb_id,
                    target_cmdb_id = EXCLUDED.target_cmdb_id,
                    source_app_name = EXCLUDED.source_app_name,
                    target_app_name = EXCLUDED.target_app_name,
                    source_account_name = EXCLUDED.source_account_name,
                    target_account_name = EXCLUDED.target_account_name,
                    source_endpoint = EXCLUDED.source_endpoint,
                    target_endpoint = EXCLUDED.target_endpoint,
                    status = EXCLUDED.status,
                    raw_fields = EXCLUDED.raw_fields,
                    updated_at = NOW()
            """

            # Pre-populate all expected keys (psycopg needs them all)
            ALL_KEYS = [
                "integration_platform", "interface_name", "source_row_hash", "source_row_num",
                "source_cmdb_id", "target_cmdb_id", "source_app_name", "target_app_name",
                "source_account_name", "target_account_name",
                "source_endpoint", "target_endpoint", "source_version", "target_version",
                "source_dc", "target_dc",
                "source_application_type", "target_application_type",
                "source_connection_type", "target_connection_type",
                "source_authentication", "target_authentication",
                "interface_owner", "source_owner", "target_owner",
                "s_team_publicmail", "t_team_publicmail",
                "s_application_linemanager", "t_application_linemanager",
                "developer",
                "frequency", "schedule", "status", "business_area",
                "interface_description", "location",
                "api_name", "topic_name", "instance", "api_postman_url", "api_spec",
                "api_payload_size", "source_payload_size", "target_payload_size",
                "data_mapping_file", "base", "git_project", "version", "tag",
                "raw_fields",
            ]
            batch: list[dict[str, Any]] = []
            for r in all_rows:
                record = {k: r.get(k) for k in ALL_KEYS}
                # raw_fields → JSON string
                if record["raw_fields"] is not None:
                    record["raw_fields"] = json.dumps(
                        record["raw_fields"], ensure_ascii=False,
                    )
                batch.append(record)

            # Bulk insert in chunks
            CHUNK = 500
            total = 0
            for i in range(0, len(batch), CHUNK):
                chunk = batch[i:i + CHUNK]
                cur.executemany(insert_sql, chunk)
                total += len(chunk)
                if i % 5000 == 0:
                    logger.info("  inserted %d/%d", total, len(batch))
            stats["rows_ingested"] = total
            pg.commit()
            logger.info("inserted %d rows", total)

        # Pass 5: trigram fallback for remaining unresolved names
        logger.info("running trigram fallback for unresolved names...")
        with pg.cursor() as cur:
            for side in ("source", "target"):
                cur.execute(TRIGRAM_FILL_SQL.format(side=side))
                logger.info("  %s trigram: %d rows updated", side, cur.rowcount)
                stats[f"trigram_{side}_filled"] = cur.rowcount
            pg.commit()

        # Summary
        with pg.cursor() as cur:
            cur.execute("""
                SELECT
                    count(*) AS total,
                    count(*) FILTER (WHERE source_cmdb_id IS NOT NULL) AS src_filled,
                    count(*) FILTER (WHERE target_cmdb_id IS NOT NULL) AS tgt_filled,
                    count(*) FILTER (WHERE source_cmdb_id IS NOT NULL
                                       AND target_cmdb_id IS NOT NULL) AS both_filled
                FROM northstar.integration_interface
            """)
            s = cur.fetchone()
            logger.info("final coverage:")
            logger.info("  total rows:       %d", s["total"])
            logger.info("  source_cmdb_id filled: %d (%.1f%%)",
                        s["src_filled"], 100.0 * s["src_filled"] / max(1, s["total"]))
            logger.info("  target_cmdb_id filled: %d (%.1f%%)",
                        s["tgt_filled"], 100.0 * s["tgt_filled"] / max(1, s["total"]))
            logger.info("  both filled:      %d (%.1f%%)",
                        s["both_filled"], 100.0 * s["both_filled"] / max(1, s["total"]))

            cur.execute("""
                SELECT integration_platform, count(*) AS n,
                       count(*) FILTER (WHERE source_cmdb_id IS NOT NULL
                                          AND target_cmdb_id IS NOT NULL) AS linked
                FROM northstar.integration_interface
                GROUP BY integration_platform
                ORDER BY n DESC
            """)
            logger.info("per-platform coverage:")
            for r in cur.fetchall():
                logger.info("  %-25s total=%-6d linked=%-6d (%.0f%%)",
                            r["integration_platform"], r["n"], r["linked"],
                            100.0 * r["linked"] / max(1, r["n"]))

    finally:
        pg.close()

    logger.info("DONE")
    for k, v in stats.items():
        logger.info("  %-30s %d", k, v)
    return 0


if __name__ == "__main__":
    sys.exit(main())
